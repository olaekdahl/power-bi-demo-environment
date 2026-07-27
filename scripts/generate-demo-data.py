#!/usr/bin/env python3
"""
Generate the PL-300 demo file set: CSV, Excel, JSON, XML and PDF.

Everything below describes ONE coherent business (Contoso Outdoor Co, FY2024) so
the files can be related to each other in a Power BI model:

    Products    -> XML/Product_Catalog.xml          (dimension)
    Stores      -> Excel/Store_Master.xlsx          (dimension)
    Regions     -> JSON/Products_Api_Response.json  (dimension, nested envelope)
    In-store    -> CSV/MonthlySales/*.csv           (fact, 12 files -> folder combine)
    Online      -> JSON/Orders_Nested.json          (fact, nested line items)
    Summary     -> PDF/Regional_Sales_Report.pdf    (aggregate, for reconciliation)

Deterministic: fixed RNG seed, so re-running produces byte-identical data. That
matters because Terraform hashes these files to decide whether to re-upload.

Usage:  python generate-demo-data.py --out ../demo-data
"""

import argparse
import csv
import datetime as dt
import json
import random
import shutil
import xml.dom.minidom
import xml.etree.ElementTree as ET
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table as PdfTable, TableStyle

SEED = 20300
YEAR = 2024

# --------------------------------------------------------------------------- #
# Reference data
# --------------------------------------------------------------------------- #

# Targets are set deliberately close to actuals so two regions beat plan and two
# miss it. A demo where every region is 40% under target teaches nothing about
# conditional formatting or KPI visuals.
REGIONS = [
    {"RegionID": 1, "RegionName": "North",   "Manager": "Dana Whitfield", "TargetRevenue": 1_600_000},
    {"RegionID": 2, "RegionName": "South",   "Manager": "Marcus Iyer",    "TargetRevenue": 1_750_000},
    {"RegionID": 3, "RegionName": "East",    "Manager": "Priya Raman",    "TargetRevenue": 1_620_000},
    {"RegionID": 4, "RegionName": "West",    "Manager": "Tomas Lindqvist","TargetRevenue": 1_800_000},
]

PRODUCTS = [
    # ID, Name, Category, Subcategory, Cost, ListPrice
    (101, "Trailblazer 2P Tent",        "Camping",  "Tents",        180.00, 379.99),
    (102, "Summit 4P Dome Tent",        "Camping",  "Tents",        265.00, 549.99),
    (103, "Alpine 20F Sleeping Bag",    "Camping",  "Sleep",         72.00, 149.99),
    (104, "Featherlite Sleeping Pad",   "Camping",  "Sleep",         38.00,  89.99),
    (105, "Cascade 65L Backpack",       "Hiking",   "Packs",        124.00, 249.99),
    (106, "Daybreak 28L Daypack",       "Hiking",   "Packs",         41.00,  94.99),
    (107, "Ridgeline Trekking Poles",   "Hiking",   "Accessories",   34.00,  79.99),
    (108, "Stormshell Rain Jacket",     "Apparel",  "Outerwear",     88.00, 199.99),
    (109, "Basecamp Down Parka",        "Apparel",  "Outerwear",    142.00, 329.99),
    (110, "Merino Trail Socks 3pk",     "Apparel",  "Socks",         14.00,  34.99),
    (111, "Kettle Camp Stove",          "Cooking",  "Stoves",        56.00, 119.99),
    (112, "Titanium Mess Kit",          "Cooking",  "Cookware",      29.00,  64.99),
]

STORES = [
    # ID, Name, City, State, RegionID, OpenDate, SqFt
    (1, "Contoso Bellevue",     "Bellevue",     "WA", 4, "2016-04-12", 12500),
    (2, "Contoso Portland",     "Portland",     "OR", 4, "2018-09-01",  9800),
    (3, "Contoso Denver",       "Denver",       "CO", 1, "2015-06-20", 14200),
    (4, "Contoso Minneapolis",  "Minneapolis",  "MN", 1, "2019-11-08",  8600),
    (5, "Contoso Austin",       "Austin",       "TX", 2, "2017-03-15", 11300),
    (6, "Contoso Atlanta",      "Atlanta",      "GA", 2, "2020-02-28",  7900),
    (7, "Contoso Boston",       "Boston",       "MA", 3, "2014-08-05", 13100),
    (8, "Contoso Philadelphia", "Philadelphia", "PA", 3, "2021-05-17",  8100),
]

# Real coordinates for the store cities, so map visuals land in the right place.
# Keep in step with scripts/sql/create-spatial-demo.sql, which builds the same
# geography points and region polygons inside SQL Server.
STORE_COORDS = {
    1: (47.6104, -122.2007),   # Bellevue, WA
    2: (45.5152, -122.6784),   # Portland, OR
    3: (39.7392, -104.9903),   # Denver, CO
    4: (44.9778, -93.2650),    # Minneapolis, MN
    5: (30.2672, -97.7431),    # Austin, TX
    6: (33.7490, -84.3880),    # Atlanta, GA
    7: (42.3601, -71.0589),    # Boston, MA
    8: (39.9526, -75.1652),    # Philadelphia, PA
}

# Approximate bounding boxes per sales region. Each box contains exactly the two
# stores assigned to that region, so point-in-polygon demos give a clean answer.
REGION_BBOX = {
    "West":  {"lon": (-125.0, -115.0), "lat": (32.0, 49.5)},
    "North": {"lon": (-115.0, -87.0),  "lat": (38.5, 49.5)},
    "South": {"lon": (-107.0, -75.0),  "lat": (24.0, 38.5)},
    "East":  {"lon": (-80.5, -66.0),   "lat": (38.5, 48.0)},
}

CHANNELS = ["In-Store", "Curbside", "Phone"]
MONTH_NAMES = ["January", "February", "March", "April", "May", "June",
               "July", "August", "September", "October", "November", "December"]

# Outdoor retail is seasonal - spring/summer peak. Used to shape the fact table
# so the demos show a real trend instead of noise.
SEASONALITY = [0.62, 0.68, 0.92, 1.08, 1.28, 1.42, 1.38, 1.22, 1.05, 0.88, 0.95, 1.15]

CUSTOMER_FIRST = ["Avery", "Bianca", "Caleb", "Dalia", "Emmett", "Farrah", "Gideon", "Hana",
                  "Ivan", "Jolene", "Kwame", "Lena", "Mateo", "Nadia", "Omar", "Petra",
                  "Quinn", "Rosa", "Sven", "Tamsin", "Ulises", "Vera", "Wes", "Yara"]
CUSTOMER_LAST = ["Alvarez", "Bhatt", "Castellano", "Duarte", "Eriksen", "Fontaine", "Gallo",
                 "Haddad", "Ishikawa", "Jankowski", "Kovacs", "Laurent", "Moreau", "Nkemdirim",
                 "Okonkwo", "Petrov", "Quintero", "Rasmussen", "Sandoval", "Tanaka"]


# --------------------------------------------------------------------------- #
# Fact generation
# --------------------------------------------------------------------------- #

def money(x):
    return round(x + 1e-9, 2)


def generate_sales(rng):
    """In-store transactions for the whole year, one row per line item."""
    rows = []
    txn_id = 500_000
    product_by_id = {p[0]: p for p in PRODUCTS}
    store_region = {s[0]: s[4] for s in STORES}

    day = dt.date(YEAR, 1, 1)
    end = dt.date(YEAR, 12, 31)
    while day <= end:
        season = SEASONALITY[day.month - 1]
        # Weekends are busier in outdoor retail.
        dow_factor = 1.45 if day.weekday() >= 5 else 1.0
        for store_id, *_ in STORES:
            base = rng.uniform(3.0, 7.0)
            n_txn = max(1, int(base * season * dow_factor))
            for _ in range(n_txn):
                txn_id += 1
                n_lines = rng.choices([1, 2, 3, 4], weights=[52, 28, 14, 6])[0]
                cust = f"C{rng.randint(1000, 4999)}"
                channel = rng.choices(CHANNELS, weights=[80, 14, 6])[0]
                for line in range(1, n_lines + 1):
                    pid = rng.choice(list(product_by_id))
                    _, name, cat, subcat, cost, price = product_by_id[pid]
                    qty = rng.choices([1, 2, 3], weights=[76, 19, 5])[0]
                    # Occasional promo discounts, heavier in clearance months.
                    disc = rng.choices([0.0, 0.10, 0.15, 0.25],
                                       weights=[70, 15, 10, 5])[0]
                    gross = price * qty
                    net = money(gross * (1 - disc))
                    rows.append({
                        "TransactionID": txn_id,
                        "LineNumber": line,
                        "OrderDate": day.isoformat(),
                        "StoreID": store_id,
                        "RegionID": store_region[store_id],
                        "ProductID": pid,
                        "Channel": channel,
                        "CustomerID": cust,
                        "Quantity": qty,
                        "UnitPrice": money(price),
                        "DiscountPct": disc,
                        "SalesAmount": net,
                        "UnitCost": money(cost),
                        "TotalCost": money(cost * qty),
                    })
        day += dt.timedelta(days=1)
    return rows


def generate_online_orders(rng, n=900):
    """Online orders, shaped for the JSON nested-expand demo."""
    orders = []
    product_by_id = {p[0]: p for p in PRODUCTS}
    for i in range(n):
        oid = f"WEB-{YEAR}-{10000 + i}"
        month = rng.choices(range(1, 13), weights=SEASONALITY)[0]
        day = rng.randint(1, 28)
        order_date = dt.date(YEAR, month, day)
        region = rng.choice(REGIONS)
        n_lines = rng.choices([1, 2, 3, 4, 5], weights=[40, 28, 18, 9, 5])[0]
        items = []
        for line in range(1, n_lines + 1):
            pid = rng.choice(list(product_by_id))
            _, name, cat, subcat, cost, price = product_by_id[pid]
            qty = rng.choices([1, 2, 3], weights=[74, 20, 6])[0]
            disc = rng.choices([0.0, 0.10, 0.20], weights=[74, 18, 8])[0]
            items.append({
                "lineNumber": line,
                "productId": pid,
                "productName": name,
                "quantity": qty,
                "unitPrice": money(price),
                "discountPct": disc,
                "lineTotal": money(price * qty * (1 - disc)),
            })
        ship = money(rng.choice([0.0, 0.0, 7.95, 12.50, 19.95]))
        orders.append({
            "orderId": oid,
            "orderDate": order_date.isoformat(),
            "status": rng.choices(["Shipped", "Delivered", "Processing", "Cancelled"],
                                  weights=[30, 60, 7, 3])[0],
            "customer": {
                "customerId": f"C{rng.randint(1000, 4999)}",
                "name": f"{rng.choice(CUSTOMER_FIRST)} {rng.choice(CUSTOMER_LAST)}",
                "email": None,
                "loyaltyTier": rng.choices(["None", "Silver", "Gold", "Platinum"],
                                           weights=[46, 30, 18, 6])[0],
            },
            "shipTo": {
                "regionId": region["RegionID"],
                "regionName": region["RegionName"],
                "postalCode": f"{rng.randint(10000, 99999)}",
            },
            "shipping": {"method": rng.choice(["Ground", "Two-Day", "Overnight"]), "cost": ship},
            "lineItems": items,
            "orderTotal": money(sum(li["lineTotal"] for li in items) + ship),
        })
    # email is filled for most but not all customers - gives a null-handling demo
    for o in orders:
        c = o["customer"]
        if rng.random() > 0.12:
            first, last = c["name"].split(" ", 1)
            c["email"] = f"{first.lower()}.{last.lower().replace(' ', '')}@example.com"
    return orders


# --------------------------------------------------------------------------- #
# CSV
# --------------------------------------------------------------------------- #

CSV_FIELDS = ["TransactionID", "LineNumber", "OrderDate", "StoreID", "RegionID", "ProductID",
              "Channel", "CustomerID", "Quantity", "UnitPrice", "DiscountPct", "SalesAmount",
              "UnitCost", "TotalCost"]


def write_clean_csv(rows, path):
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=CSV_FIELDS)
        w.writeheader()
        w.writerows(rows)


def write_monthly_csvs(rows, outdir):
    """12 files with identical schema -> the 'Combine Files from Folder' demo."""
    outdir.mkdir(parents=True, exist_ok=True)
    for m in range(1, 13):
        subset = [r for r in rows if int(r["OrderDate"][5:7]) == m]
        p = outdir / f"Sales_{YEAR}-{m:02d}.csv"
        write_clean_csv(subset, p)
    return 12


def write_messy_csv(rows, path, rng):
    """
    Deliberately dirty extract for the Power Query cleanup demo. Contains, in order:
      - 3 preamble junk lines before the header   -> Remove Top Rows
      - mixed date formats                        -> locale-aware date parsing
      - currency symbols and thousands separators -> Replace Values / typed columns
      - inconsistent region casing and whitespace -> Trim / Clean / Format
      - 'N/A', 'NULL', empty strings for nulls    -> Replace Values
      - a repeated header row in the middle       -> Filter Rows
      - a handful of exact duplicate rows         -> Remove Duplicates
      - blank lines                               -> Remove Blank Rows
      - a grand-total row at the bottom           -> Remove Bottom Rows
    """
    region_name = {r["RegionID"]: r["RegionName"] for r in REGIONS}
    sample = rows[:600]
    out = []

    out.append(["Contoso Outdoor Co - Point of Sale Extract"])
    out.append([f"Generated: {YEAR}-12-31 23:59  |  Source: POS-LEGACY-02"])
    out.append([])
    header = ["Txn ID", "Order Date", "Store", "Region", "Product ID",
              "Qty", "Unit Price", "Discount", "Sales Amount", "Notes"]
    out.append(header)

    def fmt_date(iso, style):
        d = dt.date.fromisoformat(iso)
        if style == 0:
            return d.isoformat()
        if style == 1:
            return f"{d.month:02d}/{d.day:02d}/{d.year}"
        return d.strftime("%d-%b-%Y")

    def fmt_region(name, style):
        if style == 0:
            return name
        if style == 1:
            return name.upper()
        if style == 2:
            return name.lower()
        return f"  {name} "

    made = []
    for i, r in enumerate(sample):
        rec = [
            str(r["TransactionID"]),
            fmt_date(r["OrderDate"], i % 3),
            f"Store {r['StoreID']}" if i % 5 else f" Store {r['StoreID']}  ",
            fmt_region(region_name[r["RegionID"]], i % 4),
            str(r["ProductID"]),
            str(r["Quantity"]),
            f"${r['UnitPrice']:,.2f}",
            f"{r['DiscountPct']:.0%}" if r["DiscountPct"] else rng.choice(["0%", "", "N/A"]),
            f"${r['SalesAmount']:,.2f}",
            rng.choice(["", "", "", "NULL", "N/A", "checked", "manager override"]),
        ]
        made.append(rec)

    for i, rec in enumerate(made):
        out.append(rec)
        if i == 199:
            out.append([])                 # stray blank line
            out.append(header)             # repeated header mid-file
        if i in (57, 158, 301, 455):
            out.append(list(rec))          # exact duplicate
        if i == 380:
            out.append([])

    total = sum(r["SalesAmount"] for r in sample)
    out.append([])
    out.append(["GRAND TOTAL", "", "", "", "", "", "", "", f"${total:,.2f}", ""])
    out.append(["*** End of extract - do not edit below this line ***"])

    with path.open("w", newline="", encoding="utf-8") as f:
        csv.writer(f).writerows(out)


# --------------------------------------------------------------------------- #
# Excel
# --------------------------------------------------------------------------- #

HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(color="FFFFFF", bold=True)


def _style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center")


def _autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_store_master(path):
    """Clean dimension table, exposed as a real Excel Table object."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Stores"
    cols = ["StoreID", "StoreName", "City", "State", "RegionID", "OpenDate", "SquareFeet"]
    ws.append(cols)
    for s in STORES:
        ws.append([s[0], s[1], s[2], s[3], s[4], dt.date.fromisoformat(s[5]), s[6]])
    for r in range(2, len(STORES) + 2):
        ws.cell(row=r, column=6).number_format = "yyyy-mm-dd"
    _style_header(ws, len(cols))
    _autosize(ws, [9, 22, 15, 8, 10, 13, 12])
    tbl = Table(displayName="StoreMaster", ref=f"A1:G{len(STORES) + 1}")
    tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(tbl)

    ws2 = wb.create_sheet("Regions")
    ws2.append(["RegionID", "RegionName", "RegionManager", "TargetRevenue"])
    for r in REGIONS:
        ws2.append([r["RegionID"], r["RegionName"], r["Manager"], r["TargetRevenue"]])
    for r in range(2, len(REGIONS) + 2):
        ws2.cell(row=r, column=4).number_format = '#,##0'
    _style_header(ws2, 4)
    _autosize(ws2, [10, 14, 20, 16])

    wb.save(path)


def write_sales_analysis(rows, path):
    """
    Three-sheet workbook, each sheet teaching something different:
      SalesData  - a proper Excel Table (connector shows Tables vs Sheets)
      CrossTab   - Region x Month matrix with a title block -> Unpivot demo
      Targets    - small lookup table for variance measures
      ReadMe     - a junk sheet, so the navigator has noise in it
    """
    wb = Workbook()

    # --- SalesData: aggregated to product x month so the file stays small ----
    ws = wb.active
    ws.title = "SalesData"
    agg = {}
    for r in rows:
        key = (r["ProductID"], int(r["OrderDate"][5:7]))
        a = agg.setdefault(key, {"Quantity": 0, "SalesAmount": 0.0, "TotalCost": 0.0})
        a["Quantity"] += r["Quantity"]
        a["SalesAmount"] += r["SalesAmount"]
        a["TotalCost"] += r["TotalCost"]
    prod = {p[0]: p for p in PRODUCTS}
    cols = ["ProductID", "ProductName", "Category", "Subcategory", "MonthNumber",
            "MonthName", "Quantity", "SalesAmount", "TotalCost"]
    ws.append(cols)
    n = 0
    for (pid, m) in sorted(agg):
        a = agg[(pid, m)]
        p = prod[pid]
        ws.append([pid, p[1], p[2], p[3], m, MONTH_NAMES[m - 1],
                   a["Quantity"], money(a["SalesAmount"]), money(a["TotalCost"])])
        n += 1
    for r in range(2, n + 2):
        for c in (8, 9):
            ws.cell(row=r, column=c).number_format = '#,##0.00'
    _style_header(ws, len(cols))
    _autosize(ws, [11, 26, 12, 14, 13, 12, 10, 14, 12])
    t = Table(displayName="SalesData", ref=f"A1:I{n + 1}")
    t.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws.add_table(t)

    # --- CrossTab: the classic unpivot candidate ----------------------------
    ws2 = wb.create_sheet("CrossTab")
    ws2["A1"] = "Contoso Outdoor Co"
    ws2["A1"].font = Font(bold=True, size=14)
    ws2["A2"] = f"Net Revenue by Region and Month - FY{YEAR}"
    ws2["A2"].font = Font(italic=True, size=11)
    ws2["A3"] = "(figures in USD, excludes online channel)"
    ws2["A3"].font = Font(italic=True, size=9, color="808080")
    # row 4 blank, row 5 is the real header
    matrix = {}
    region_of_store = {s[0]: s[4] for s in STORES}
    for r in rows:
        rid = region_of_store[r["StoreID"]]
        matrix.setdefault(rid, [0.0] * 12)[int(r["OrderDate"][5:7]) - 1] += r["SalesAmount"]
    ws2.append([])  # row 4
    ws2.append(["Region"] + MONTH_NAMES)  # row 5
    for reg in REGIONS:
        vals = matrix.get(reg["RegionID"], [0.0] * 12)
        ws2.append([reg["RegionName"]] + [money(v) for v in vals])
    hdr_row = 5
    _style_header(ws2, 13, row=hdr_row)
    for r in range(hdr_row + 1, hdr_row + 1 + len(REGIONS)):
        for c in range(2, 14):
            ws2.cell(row=r, column=c).number_format = '#,##0'
    _autosize(ws2, [14] + [12] * 12)

    # --- Targets -----------------------------------------------------------
    ws3 = wb.create_sheet("Targets")
    ws3.append(["RegionID", "RegionName", "Quarter", "RevenueTarget"])
    for reg in REGIONS:
        for q in range(1, 5):
            share = [0.20, 0.30, 0.30, 0.20][q - 1]
            ws3.append([reg["RegionID"], reg["RegionName"], f"Q{q}",
                        round(reg["TargetRevenue"] * share)])
    for r in range(2, len(REGIONS) * 4 + 2):
        ws3.cell(row=r, column=4).number_format = '#,##0'
    _style_header(ws3, 4)
    _autosize(ws3, [10, 14, 10, 16])

    # --- ReadMe: intentional noise in the navigator -------------------------
    ws4 = wb.create_sheet("ReadMe")
    ws4["A1"] = "Internal use only."
    ws4["A3"] = "SalesData  - transaction extract aggregated to product/month."
    ws4["A4"] = "CrossTab   - formatted for the monthly business review deck."
    ws4["A5"] = "Targets    - from Finance, quarterly plan."
    ws4["A7"] = "Contact: bi-team@contoso.example"
    _autosize(ws4, [70])

    wb.save(path)


# --------------------------------------------------------------------------- #
# JSON
# --------------------------------------------------------------------------- #

def write_orders_json(orders, path):
    with path.open("w", encoding="utf-8") as f:
        json.dump(orders, f, indent=2)


def write_products_api_json(path):
    """
    API-shaped envelope: metadata wrapper + nested arrays. Teaches drilling from
    a record into a list, then converting to a table.
    """
    payload = {
        "apiVersion": "2024-11-01",
        "endpoint": "/v2/catalog/products",
        "generatedAt": f"{YEAR}-12-31T23:59:00Z",
        "pagination": {"page": 1, "pageSize": 100, "totalRecords": len(PRODUCTS), "hasMore": False},
        "data": [
            {
                "productId": p[0],
                "productName": p[1],
                "classification": {"category": p[2], "subcategory": p[3]},
                "pricing": {"standardCost": p[4], "listPrice": p[5],
                            "marginPct": round((p[5] - p[4]) / p[5], 4)},
                "attributes": {
                    "sku": f"CO-{p[2][:3].upper()}-{p[0]}",
                    "active": True,
                    "tags": sorted({p[2].lower(), p[3].lower(),
                                    "bestseller" if p[0] in (105, 108, 103) else "standard"}),
                },
            }
            for p in PRODUCTS
        ],
        "regions": [
            {"regionId": r["RegionID"], "regionName": r["RegionName"],
             "manager": r["Manager"], "targetRevenue": r["TargetRevenue"]}
            for r in REGIONS
        ],
    }
    with path.open("w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2)


# --------------------------------------------------------------------------- #
# XML
# --------------------------------------------------------------------------- #

def _pretty_xml(root, path):
    raw = ET.tostring(root, encoding="utf-8")
    parsed = xml.dom.minidom.parseString(raw)
    with path.open("w", encoding="utf-8") as f:
        f.write(parsed.toprettyxml(indent="  "))


def write_product_catalog_xml(path):
    """Hierarchical: Catalog > Category > Product, with attributes AND elements."""
    root = ET.Element("ProductCatalog", {"company": "Contoso Outdoor Co",
                                         "fiscalYear": str(YEAR),
                                         "currency": "USD"})
    by_cat = {}
    for p in PRODUCTS:
        by_cat.setdefault(p[2], []).append(p)
    for cat, prods in by_cat.items():
        cnode = ET.SubElement(root, "Category", {"name": cat, "productCount": str(len(prods))})
        for p in prods:
            pnode = ET.SubElement(cnode, "Product", {"id": str(p[0]),
                                                     "sku": f"CO-{p[2][:3].upper()}-{p[0]}"})
            ET.SubElement(pnode, "Name").text = p[1]
            ET.SubElement(pnode, "Subcategory").text = p[3]
            ET.SubElement(pnode, "StandardCost").text = f"{p[4]:.2f}"
            ET.SubElement(pnode, "ListPrice").text = f"{p[5]:.2f}"
            ET.SubElement(pnode, "MarginPct").text = f"{(p[5] - p[4]) / p[5]:.4f}"
            spec = ET.SubElement(pnode, "Specifications")
            ET.SubElement(spec, "WeightGrams").text = str(200 + p[0] * 7)
            ET.SubElement(spec, "WarrantyMonths").text = "24" if p[2] != "Apparel" else "12"
            ET.SubElement(spec, "CountryOfOrigin").text = "Vietnam" if p[0] % 2 else "China"
    _pretty_xml(root, path)


def write_employees_xml(path, rng):
    """Flatter XML with a repeating element - easy second XML example."""
    root = ET.Element("Employees", {"effectiveDate": f"{YEAR}-12-31"})
    titles = ["Store Manager", "Assistant Manager", "Sales Associate",
              "Inventory Lead", "Visual Merchandiser"]
    eid = 4000
    for s in STORES:
        for i, title in enumerate(titles):
            eid += 1
            e = ET.SubElement(root, "Employee", {"employeeId": str(eid)})
            ET.SubElement(e, "FirstName").text = rng.choice(CUSTOMER_FIRST)
            ET.SubElement(e, "LastName").text = rng.choice(CUSTOMER_LAST)
            ET.SubElement(e, "Title").text = title
            ET.SubElement(e, "StoreID").text = str(s[0])
            ET.SubElement(e, "RegionID").text = str(s[4])
            ET.SubElement(e, "HireDate").text = (
                dt.date(rng.randint(2015, 2023), rng.randint(1, 12), rng.randint(1, 28)).isoformat())
            ET.SubElement(e, "AnnualSalary").text = str(
                {0: 78000, 1: 61000, 2: 44000, 3: 52000, 4: 49000}[i] + rng.randint(0, 6) * 500)
            ET.SubElement(e, "FullTime").text = "true" if i < 4 or rng.random() > 0.4 else "false"
    _pretty_xml(root, path)


# --------------------------------------------------------------------------- #
# PDF
# --------------------------------------------------------------------------- #

PDF_TABLE_STYLE = TableStyle([
    ("BACKGROUND",    (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
    ("TEXTCOLOR",     (0, 0), (-1, 0), colors.white),
    ("FONTNAME",      (0, 0), (-1, 0), "Helvetica-Bold"),
    ("FONTSIZE",      (0, 0), (-1, -1), 9),
    ("ALIGN",         (1, 0), (-1, -1), "RIGHT"),
    ("ALIGN",         (0, 0), (0, -1), "LEFT"),
    ("GRID",          (0, 0), (-1, -1), 0.4, colors.HexColor("#9DC3E6")),
    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EAF2FA")]),
    ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ("TOPPADDING",    (0, 0), (-1, -1), 5),
])


def write_regional_pdf(rows, orders, path):
    """
    Ruled tables so the Power BI PDF connector detects them cleanly.
    Page 1: region x quarter. Page 2: product performance. Page 3: online channel.
    """
    styles = getSampleStyleSheet()
    doc = SimpleDocTemplate(str(path), pagesize=letter,
                            leftMargin=0.7 * inch, rightMargin=0.7 * inch,
                            topMargin=0.7 * inch, bottomMargin=0.7 * inch,
                            title=f"Contoso Outdoor Co - Regional Sales Report FY{YEAR}",
                            author="Contoso BI Team")
    story = []
    region_of_store = {s[0]: s[4] for s in STORES}
    rname = {r["RegionID"]: r["RegionName"] for r in REGIONS}

    story.append(Paragraph(f"<b>Contoso Outdoor Co</b>", styles["Title"]))
    story.append(Paragraph(f"Regional Sales Report &mdash; Fiscal Year {YEAR}", styles["Heading2"]))
    story.append(Paragraph("Prepared by the BI Team. In-store channel only unless noted.",
                           styles["Normal"]))
    story.append(Spacer(1, 0.22 * inch))

    # --- Table 1: region x quarter ----------------------------------------
    story.append(Paragraph("<b>Table 1. Net Revenue by Region and Quarter</b>", styles["Heading3"]))
    story.append(Spacer(1, 0.08 * inch))
    q_mat = {r["RegionID"]: [0.0] * 4 for r in REGIONS}
    for r in rows:
        q = (int(r["OrderDate"][5:7]) - 1) // 3
        q_mat[region_of_store[r["StoreID"]]][q] += r["SalesAmount"]
    data = [["Region", "Q1", "Q2", "Q3", "Q4", "Full Year", "Target", "Var %"]]
    for reg in REGIONS:
        v = q_mat[reg["RegionID"]]
        fy = sum(v)
        tgt = reg["TargetRevenue"]
        data.append([reg["RegionName"]] + [f"{x:,.0f}" for x in v] +
                    [f"{fy:,.0f}", f"{tgt:,.0f}", f"{(fy / tgt - 1) * 100:+.1f}%"])
    tot = [sum(q_mat[r["RegionID"]][q] for r in REGIONS) for q in range(4)]
    tot_fy = sum(tot)
    tot_tgt = sum(r["TargetRevenue"] for r in REGIONS)
    data.append(["TOTAL"] + [f"{x:,.0f}" for x in tot] +
                [f"{tot_fy:,.0f}", f"{tot_tgt:,.0f}", f"{(tot_fy / tot_tgt - 1) * 100:+.1f}%"])
    t = PdfTable(data, colWidths=[1.15 * inch] + [0.86 * inch] * 7, repeatRows=1)
    st = TableStyle(PDF_TABLE_STYLE.getCommands())
    st.add("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold")
    st.add("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#D6E4F0"))
    t.setStyle(st)
    story.append(t)
    story.append(Spacer(1, 0.28 * inch))

    # --- Table 2: store detail --------------------------------------------
    story.append(Paragraph("<b>Table 2. Store Performance Detail</b>", styles["Heading3"]))
    story.append(Spacer(1, 0.08 * inch))
    s_agg = {}
    for r in rows:
        a = s_agg.setdefault(r["StoreID"], {"rev": 0.0, "cost": 0.0, "units": 0, "txn": set()})
        a["rev"] += r["SalesAmount"]
        a["cost"] += r["TotalCost"]
        a["units"] += r["Quantity"]
        a["txn"].add(r["TransactionID"])
    data2 = [["Store", "Region", "Revenue", "Cost", "Margin", "Margin %", "Units", "Transactions"]]
    for s in STORES:
        a = s_agg[s[0]]
        margin = a["rev"] - a["cost"]
        data2.append([s[1].replace("Contoso ", ""), rname[s[4]],
                      f"{a['rev']:,.0f}", f"{a['cost']:,.0f}", f"{margin:,.0f}",
                      f"{margin / a['rev'] * 100:.1f}%", f"{a['units']:,}", f"{len(a['txn']):,}"])
    t2 = PdfTable(data2, colWidths=[1.15 * inch, 0.75 * inch, 0.95 * inch, 0.9 * inch,
                                   0.9 * inch, 0.75 * inch, 0.7 * inch, 0.95 * inch],
                  repeatRows=1)
    t2.setStyle(PDF_TABLE_STYLE)
    story.append(t2)
    story.append(PageBreak())

    # --- Table 3: product performance -------------------------------------
    story.append(Paragraph("<b>Table 3. Product Performance</b>", styles["Heading3"]))
    story.append(Spacer(1, 0.08 * inch))
    p_agg = {}
    for r in rows:
        a = p_agg.setdefault(r["ProductID"], {"rev": 0.0, "units": 0})
        a["rev"] += r["SalesAmount"]
        a["units"] += r["Quantity"]
    data3 = [["Product", "Category", "Subcategory", "Units", "Revenue", "Avg Price"]]
    for p in sorted(PRODUCTS, key=lambda x: -p_agg[x[0]]["rev"]):
        a = p_agg[p[0]]
        data3.append([p[1], p[2], p[3], f"{a['units']:,}", f"{a['rev']:,.0f}",
                      f"{a['rev'] / a['units']:,.2f}"])
    t3 = PdfTable(data3, colWidths=[2.0 * inch, 0.95 * inch, 1.1 * inch,
                                    0.75 * inch, 1.0 * inch, 0.9 * inch], repeatRows=1)
    t3.setStyle(PDF_TABLE_STYLE)
    story.append(t3)
    story.append(Spacer(1, 0.28 * inch))

    # --- Table 4: online channel ------------------------------------------
    story.append(Paragraph("<b>Table 4. Online Channel by Loyalty Tier</b>", styles["Heading3"]))
    story.append(Spacer(1, 0.08 * inch))
    l_agg = {}
    for o in orders:
        if o["status"] == "Cancelled":
            continue
        tier = o["customer"]["loyaltyTier"]
        a = l_agg.setdefault(tier, {"rev": 0.0, "orders": 0, "items": 0})
        a["rev"] += o["orderTotal"]
        a["orders"] += 1
        a["items"] += sum(li["quantity"] for li in o["lineItems"])
    data4 = [["Loyalty Tier", "Orders", "Items", "Revenue", "Avg Order Value"]]
    for tier in ["None", "Silver", "Gold", "Platinum"]:
        a = l_agg.get(tier)
        if not a:
            continue
        data4.append([tier, f"{a['orders']:,}", f"{a['items']:,}",
                      f"{a['rev']:,.0f}", f"{a['rev'] / a['orders']:,.2f}"])
    t4 = PdfTable(data4, colWidths=[1.3 * inch, 0.9 * inch, 0.9 * inch, 1.1 * inch, 1.3 * inch],
                  repeatRows=1)
    t4.setStyle(PDF_TABLE_STYLE)
    story.append(t4)
    story.append(Spacer(1, 0.3 * inch))
    story.append(Paragraph(
        "<i>Note: figures in this report are generated sample data for training purposes "
        "(Microsoft PL-300). They reconcile with the CSV, Excel, JSON and XML files in the "
        "same demo data set.</i>", styles["Normal"]))

    doc.build(story)


def write_quarterly_pdf(rows, path):
    """Single-table PDF - the simplest possible PDF connector demo."""
    styles = getSampleStyleSheet()
    doc = SimpleDocTemplate(str(path), pagesize=letter,
                            leftMargin=0.8 * inch, rightMargin=0.8 * inch,
                            topMargin=0.8 * inch, bottomMargin=0.8 * inch,
                            title=f"Contoso Outdoor Co - Quarterly Summary FY{YEAR}")
    story = [Paragraph("<b>Contoso Outdoor Co</b>", styles["Title"]),
             Paragraph(f"Quarterly Summary &mdash; FY{YEAR}", styles["Heading2"]),
             Spacer(1, 0.25 * inch)]
    m_agg = {m: {"rev": 0.0, "cost": 0.0, "units": 0} for m in range(1, 13)}
    for r in rows:
        m = int(r["OrderDate"][5:7])
        m_agg[m]["rev"] += r["SalesAmount"]
        m_agg[m]["cost"] += r["TotalCost"]
        m_agg[m]["units"] += r["Quantity"]
    data = [["Month", "Quarter", "Units Sold", "Revenue", "Cost of Goods", "Gross Margin"]]
    for m in range(1, 13):
        a = m_agg[m]
        data.append([MONTH_NAMES[m - 1], f"Q{(m - 1) // 3 + 1}", f"{a['units']:,}",
                     f"{a['rev']:,.0f}", f"{a['cost']:,.0f}", f"{a['rev'] - a['cost']:,.0f}"])
    t = PdfTable(data, colWidths=[1.2 * inch, 0.85 * inch, 1.05 * inch,
                                  1.15 * inch, 1.25 * inch, 1.25 * inch], repeatRows=1)
    t.setStyle(PDF_TABLE_STYLE)
    story.append(t)
    doc.build(story)


# --------------------------------------------------------------------------- #
# Spatial
# --------------------------------------------------------------------------- #

def _ccw_ring(lon_range, lat_range):
    """
    Counter-clockwise exterior ring for a bounding box.

    Orientation is not cosmetic. SQL Server's `geography` type uses the
    left-hand rule, so a clockwise ring describes everything on the *outside* -
    you get a polygon covering nearly the whole planet instead of one US region.
    GeoJSON (RFC 7946) wants counter-clockwise exteriors too, so one order serves
    both.
    """
    lon0, lon1 = lon_range
    lat0, lat1 = lat_range
    return [[lon0, lat0], [lon1, lat0], [lon1, lat1], [lon0, lat1], [lon0, lat0]]


def write_store_locations_csv(path):
    """Lat/long plus WKT - the simplest possible map-visual source."""
    region_of = {r["RegionID"]: r["RegionName"] for r in REGIONS}
    cols = ["StoreID", "StoreName", "City", "State", "RegionID", "RegionName",
            "Latitude", "Longitude", "WKT"]
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(cols)
        for s in STORES:
            lat, lon = STORE_COORDS[s[0]]
            w.writerow([s[0], s[1], s[2], s[3], s[4], region_of[s[4]],
                        f"{lat:.4f}", f"{lon:.4f}", f"POINT ({lon:.4f} {lat:.4f})"])


def write_store_locations_geojson(path):
    region_of = {r["RegionID"]: r["RegionName"] for r in REGIONS}
    features = []
    for s in STORES:
        lat, lon = STORE_COORDS[s[0]]
        features.append({
            "type": "Feature",
            "geometry": {"type": "Point", "coordinates": [lon, lat]},
            "properties": {
                "StoreID": s[0], "StoreName": s[1], "City": s[2], "State": s[3],
                "RegionID": s[4], "RegionName": region_of[s[4]], "SquareFeet": s[6],
            },
        })
    payload = {"type": "FeatureCollection",
               "crs": {"type": "name", "properties": {"name": "urn:ogc:def:crs:OGC:1.3:CRS84"}},
               "features": features}
    with path.open("w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2)


def write_regions_geojson(path):
    features = []
    for reg in REGIONS:
        box = REGION_BBOX[reg["RegionName"]]
        features.append({
            "type": "Feature",
            "geometry": {"type": "Polygon", "coordinates": [_ccw_ring(box["lon"], box["lat"])]},
            "properties": {
                "RegionID": reg["RegionID"], "RegionName": reg["RegionName"],
                "Manager": reg["Manager"], "TargetRevenue": reg["TargetRevenue"],
            },
        })
    with path.open("w", encoding="utf-8") as f:
        json.dump({"type": "FeatureCollection", "features": features}, f, indent=2)


def write_regions_topojson(path):
    """
    Minimal TopoJSON for Power BI's Shape Map visual.

    No `transform` member, which means arc positions are absolute rather than
    delta-encoded - the spec only delta-encodes quantized topologies, and skipping
    quantization keeps this readable and hand-verifiable.
    """
    arcs = []
    geometries = []
    for i, reg in enumerate(REGIONS):
        box = REGION_BBOX[reg["RegionName"]]
        arcs.append(_ccw_ring(box["lon"], box["lat"]))
        geometries.append({
            "type": "Polygon",
            "arcs": [[i]],
            "properties": {"RegionName": reg["RegionName"], "RegionID": reg["RegionID"]},
        })
    payload = {
        "type": "Topology",
        "objects": {"regions": {"type": "GeometryCollection", "geometries": geometries}},
        "arcs": arcs,
    }
    with path.open("w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2)


def write_customer_locations_csv(path, rng, per_store=60):
    """
    Customers scattered around each store, for distance and catchment demos.
    Offsets are small enough that every customer stays inside its store's region
    polygon, which keeps the point-in-polygon results tidy.
    """
    region_of = {r["RegionID"]: r["RegionName"] for r in REGIONS}
    cols = ["CustomerID", "CustomerName", "HomeStoreID", "RegionName",
            "Latitude", "Longitude", "LoyaltyTier"]
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(cols)
        cid = 1000
        for s in STORES:
            base_lat, base_lon = STORE_COORDS[s[0]]
            for _ in range(per_store):
                cid += 1
                lat = base_lat + rng.uniform(-0.45, 0.45)
                lon = base_lon + rng.uniform(-0.55, 0.55)
                w.writerow([f"C{cid}",
                            f"{rng.choice(CUSTOMER_FIRST)} {rng.choice(CUSTOMER_LAST)}",
                            s[0], region_of[s[4]],
                            f"{lat:.5f}", f"{lon:.5f}",
                            rng.choices(["None", "Silver", "Gold", "Platinum"],
                                        weights=[46, 30, 18, 6])[0]])


def verify_spatial_consistency():
    """
    Assert each store falls inside exactly one region polygon, and that the
    polygon matches the store's assigned region. A silent mismatch here would
    make the SQL point-in-polygon demo contradict the dimension tables.
    """
    region_of = {r["RegionID"]: r["RegionName"] for r in REGIONS}
    problems = []
    for s in STORES:
        lat, lon = STORE_COORDS[s[0]]
        hits = [name for name, box in REGION_BBOX.items()
                if box["lon"][0] <= lon <= box["lon"][1] and box["lat"][0] <= lat <= box["lat"][1]]
        expected = region_of[s[4]]
        if hits != [expected]:
            problems.append(f"{s[1]} ({lat},{lon}) -> {hits}, expected exactly ['{expected}']")
    if problems:
        raise SystemExit("Spatial consistency check failed:\n  " + "\n  ".join(problems))
    return len(STORES)


# --------------------------------------------------------------------------- #
# Driver
# --------------------------------------------------------------------------- #

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", default="demo-data", help="output directory")
    args = ap.parse_args()

    out = Path(args.out).resolve()
    if out.exists():
        shutil.rmtree(out)
    for sub in ("CSV/MonthlySales", "Excel", "JSON", "XML", "PDF", "Spatial"):
        (out / sub).mkdir(parents=True, exist_ok=True)

    rng = random.Random(SEED)
    print("Generating fact data ...")
    rows = generate_sales(rng)
    orders = generate_online_orders(rng)
    print(f"  in-store line items : {len(rows):,}")
    print(f"  online orders       : {len(orders):,}")

    print("Writing CSV ...")
    write_clean_csv(rows, out / "CSV" / "Sales_Transactions.csv")
    write_messy_csv(rows, out / "CSV" / "Sales_Transactions_MESSY.csv", random.Random(SEED + 1))
    n = write_monthly_csvs(rows, out / "CSV" / "MonthlySales")
    print(f"  Sales_Transactions.csv, Sales_Transactions_MESSY.csv, MonthlySales/ ({n} files)")

    print("Writing Excel ...")
    write_store_master(out / "Excel" / "Store_Master.xlsx")
    write_sales_analysis(rows, out / "Excel" / "Product_Sales_Analysis.xlsx")
    print("  Store_Master.xlsx, Product_Sales_Analysis.xlsx")

    print("Writing JSON ...")
    write_orders_json(orders, out / "JSON" / "Orders_Nested.json")
    write_products_api_json(out / "JSON" / "Products_Api_Response.json")
    print("  Orders_Nested.json, Products_Api_Response.json")

    print("Writing XML ...")
    write_product_catalog_xml(out / "XML" / "Product_Catalog.xml")
    write_employees_xml(out / "XML" / "Employees.xml", random.Random(SEED + 2))
    print("  Product_Catalog.xml, Employees.xml")

    print("Writing spatial ...")
    n = verify_spatial_consistency()
    print(f"  {n}/{len(STORES)} stores verified inside exactly one region polygon")
    write_store_locations_csv(out / "Spatial" / "Store_Locations.csv")
    write_store_locations_geojson(out / "Spatial" / "Store_Locations.geojson")
    write_regions_geojson(out / "Spatial" / "Contoso_Regions.geojson")
    write_regions_topojson(out / "Spatial" / "Contoso_Regions.topojson")
    write_customer_locations_csv(out / "Spatial" / "Customer_Locations.csv", random.Random(SEED + 3))
    print("  Store_Locations.csv/.geojson, Contoso_Regions.geojson/.topojson, Customer_Locations.csv")

    print("Writing PDF ...")
    write_regional_pdf(rows, orders, out / "PDF" / "Regional_Sales_Report.pdf")
    write_quarterly_pdf(rows, out / "PDF" / "Quarterly_Summary.pdf")
    print("  Regional_Sales_Report.pdf, Quarterly_Summary.pdf")

    files = sorted(p for p in out.rglob("*") if p.is_file())
    total = sum(p.stat().st_size for p in files)
    print(f"\n{len(files)} files, {total / 1024 / 1024:.2f} MB total -> {out}")


if __name__ == "__main__":
    main()
